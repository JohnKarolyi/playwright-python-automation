import os
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(__file__))


def load_excel_data(path):
    full_path = os.path.join(BASE_DIR, path)

    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Excel file not found: {full_path}")

    workbook = load_workbook(full_path)
    sheet = workbook.active

    return [
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]


def get_test_data():
    return load_excel_data("data/valid_users.xlsx")


def get_invalid_test_data():
    return load_excel_data("data/invalid_users.xlsx")